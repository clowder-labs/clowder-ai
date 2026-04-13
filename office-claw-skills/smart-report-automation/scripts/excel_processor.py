#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Excel 智能报表自动化处理器
功能：数据计算、格式美化、图表生成、PDF导出
"""

import argparse
import json
import sys
from pathlib import Path
from typing import List, Dict, Any, Optional

try:
    import openpyxl
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
    from openpyxl.formatting.rule import DataBarRule, ColorScaleRule
    from openpyxl.chart import (
        BarChart, LineChart, PieChart, ScatterChart, AreaChart,
        Reference
    )
    from openpyxl.chart.label import DataLabelList
    from openpyxl.utils import get_column_letter
except ImportError:
    print("错误：缺少 openpyxl 库，请运行: pip install openpyxl>=3.1.0")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    print("错误：缺少 pandas 库，请运行: pip install pandas>=2.0.0")
    sys.exit(1)


class ExcelProcessor:
    """Excel 报表处理器"""
    
    def __init__(self, input_file: str, output_file: str):
        """
        初始化处理器
        
        Args:
            input_file: 输入 Excel 文件路径
            output_file: 输出文件路径
        """
        self.input_file = Path(input_file)
        self.output_file = Path(output_file)
        self.workbook = None
        self.sheet = None
        
    def load_workbook(self):
        """加载工作簿"""
        if not self.input_file.exists():
            raise FileNotFoundError(f"输入文件不存在: {self.input_file}")
        
        self.workbook = load_workbook(self.input_file)
        self.sheet = self.workbook.active
        
    def execute_calculations(self, calculations: List[Dict[str, Any]]) -> Dict[str, Any]:
        """
        执行数据计算
        
        Args:
            calculations: 计算任务列表
                格式: [{"type": "sum", "range": "A1:A10", "target": "B1", "sheet": "Sheet1"}]
        
        Returns:
            计算结果字典
        """
        results = {}
        
        for calc in calculations:
            calc_type = calc.get("type", "sum")
            data_range = calc.get("range", "")
            target_cell = calc.get("target", "")
            sheet_name = calc.get("sheet", self.sheet.title)
            
            # 获取工作表
            ws = self.workbook[sheet_name] if sheet_name in self.workbook.sheetnames else self.sheet
            
            # 解析范围
            if ":" in data_range:
                start_cell, end_cell = data_range.split(":")
            else:
                start_cell = end_cell = data_range
            
            # 获取数据
            data_range_obj = Reference(ws, min_col=ws[start_cell].column, min_row=ws[start_cell].row,
                                       max_col=ws[end_cell].column, max_row=ws[end_cell].row)
            values = [cell.value for row in data_range_obj.cells for cell in row if cell.value is not None]
            
            # 执行计算
            if calc_type == "sum":
                result = sum(values)
            elif calc_type == "average":
                result = sum(values) / len(values) if values else 0
            elif calc_type == "count":
                result = len(values)
            elif calc_type == "max":
                result = max(values) if values else 0
            elif calc_type == "min":
                result = min(values) if values else 0
            else:
                result = None
            
            # 写入结果
            if target_cell and result is not None:
                ws[target_cell] = result
            
            results[f"{calc_type}_{data_range}"] = result
        
        return results
    
    def apply_formatting(self, formatting: Dict[str, Any]):
        """
        应用格式化
        
        Args:
            formatting: 格式化配置
                格式参考 references/excel-format-spec.md
        """
        # 字体设置
        if "font" in formatting:
            font_config = formatting["font"]
            for range_str, font_opts in font_config.items():
                font = Font(
                    name=font_opts.get("name", "Arial"),
                    size=font_opts.get("size", 11),
                    bold=font_opts.get("bold", False),
                    italic=font_opts.get("italic", False),
                    color=font_opts.get("color", "000000")
                )
                for row in self.sheet[range_str]:
                    for cell in row:
                        cell.font = font
        
        # 边框设置
        if "border" in formatting:
            border_config = formatting["border"]
            for range_str, border_opts in border_config.items():
                border_style = Border(
                    left=Side(style=border_opts.get("left", "thin")),
                    right=Side(style=border_opts.get("right", "thin")),
                    top=Side(style=border_opts.get("top", "thin")),
                    bottom=Side(style=border_opts.get("bottom", "thin"))
                )
                for row in self.sheet[range_str]:
                    for cell in row:
                        cell.border = border_style
        
        # 填充设置
        if "fill" in formatting:
            fill_config = formatting["fill"]
            for range_str, fill_opts in fill_config.items():
                fill = PatternFill(
                    start_color=fill_opts.get("color", "FFFFFF"),
                    end_color=fill_opts.get("color", "FFFFFF"),
                    fill_type="solid"
                )
                for row in self.sheet[range_str]:
                    for cell in row:
                        cell.fill = fill
        
        # 对齐设置
        if "alignment" in formatting:
            align_config = formatting["alignment"]
            for range_str, align_opts in align_config.items():
                alignment = Alignment(
                    horizontal=align_opts.get("horizontal", "left"),
                    vertical=align_opts.get("vertical", "center"),
                    wrap_text=align_opts.get("wrap_text", False)
                )
                for row in self.sheet[range_str]:
                    for cell in row:
                        cell.alignment = alignment
        
        # 条件格式
        if "conditional_format" in formatting:
            cf_config = formatting["conditional_format"]
            for range_str, cf_opts in cf_config.items():
                cf_type = cf_opts.get("type", "data_bar")
                
                if cf_type == "data_bar":
                    rule = DataBarRule(
                        start_type=cf_opts.get("start_type", "min"),
                        end_type=cf_opts.get("end_type", "max"),
                        color=cf_opts.get("color", "638EC6")
                    )
                    self.sheet.conditional_formatting.add(range_str, rule)
                
                elif cf_type == "color_scale":
                    rule = ColorScaleRule(
                        start_type=cf_opts.get("start_type", "min"),
                        start_color=cf_opts.get("start_color", "F8696B"),
                        mid_type=cf_opts.get("mid_type", "percentile"),
                        mid_color=cf_opts.get("mid_color", "FFEB84"),
                        end_type=cf_opts.get("end_type", "max"),
                        end_color=cf_opts.get("end_color", "63BE7B")
                    )
                    self.sheet.conditional_formatting.add(range_str, rule)
    
    def create_charts(self, charts: List[Dict[str, Any]]):
        """
        创建图表
        
        Args:
            charts: 图表配置列表
                格式参考 references/chart-config-guide.md
        """
        for chart_config in charts:
            chart_type = chart_config.get("type", "bar")
            title = chart_config.get("title", "")
            data_range = chart_config.get("data_range", "")
            categories_range = chart_config.get("categories_range", "")
            position = chart_config.get("position", "E5")
            sheet_name = chart_config.get("sheet", self.sheet.title)
            
            ws = self.workbook[sheet_name] if sheet_name in self.workbook.sheetnames else self.sheet
            
            # 解析数据范围
            data_start, data_end = data_range.split(":") if ":" in data_range else (data_range, data_range)
            
            # 创建图表对象
            if chart_type == "bar":
                chart = BarChart()
            elif chart_type == "line":
                chart = LineChart()
            elif chart_type == "pie":
                chart = PieChart()
            elif chart_type == "scatter":
                chart = ScatterChart()
            elif chart_type == "area":
                chart = AreaChart()
            else:
                chart = BarChart()
            
            # 设置数据
            data_ref = Reference(ws, min_col=ws[data_start].column, min_row=ws[data_start].row,
                                max_col=ws[data_end].column, max_row=ws[data_end].row)
            
            if categories_range:
                cat_start, cat_end = categories_range.split(":") if ":" in categories_range else (categories_range, categories_range)
                cat_ref = Reference(ws, min_col=ws[cat_start].column, min_row=ws[cat_start].row,
                                   max_col=ws[cat_end].column, max_row=ws[cat_end].row)
                chart.add_data(data_ref, titles_from_data=True)
                chart.set_categories(cat_ref)
            else:
                chart.add_data(data_ref, titles_from_data=True)
            
            # 设置标题和样式
            chart.title = title
            chart.style = chart_config.get("style", 10)
            
            # 图例设置
            if chart_config.get("show_legend", True):
                chart.legend.position = chart_config.get("legend_position", "r")
            else:
                chart.legend = None
            
            # 数据标签
            if chart_config.get("show_labels", False):
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showVal = True
            
            # 添加图表到工作表
            ws.add_chart(chart, position)
    
    def save(self) -> str:
        """保存工作簿"""
        self.output_file.parent.mkdir(parents=True, exist_ok=True)
        self.workbook.save(self.output_file)
        return str(self.output_file)


def process_excel_report(
    input_file: str,
    output_file: str,
    calculations: Optional[List[Dict[str, Any]]] = None,
    formatting: Optional[Dict[str, Any]] = None,
    charts: Optional[List[Dict[str, Any]]] = None,
    export_pdf: bool = False
) -> str:
    """
    Excel 报表自动化处理主函数
    
    Args:
        input_file: 输入 Excel 文件路径
        output_file: 输出文件路径
        calculations: 计算任务列表
        formatting: 格式化配置
        charts: 图表配置列表
        export_pdf: 是否导出为 PDF
    
    Returns:
        输出文件路径
    """
    processor = ExcelProcessor(input_file, output_file)
    processor.load_workbook()
    
    # 执行计算
    if calculations:
        processor.execute_calculations(calculations)
    
    # 应用格式
    if formatting:
        processor.apply_formatting(formatting)
    
    # 创建图表
    if charts:
        processor.create_charts(charts)
    
    # 保存文件
    result_path = processor.save()
    
    # 导出 PDF（如果需要）
    if export_pdf:
        try:
            from format_converter import convert_to_pdf
            pdf_path = convert_to_pdf(result_path, output_format="pdf")
            print(f"PDF 已导出: {pdf_path}")
        except Exception as e:
            print(f"警告：PDF 导出失败: {str(e)}")
    
    return result_path


def main():
    """命令行入口"""
    parser = argparse.ArgumentParser(description="Excel 智能报表自动化处理器")
    parser.add_argument("--input", required=True, help="输入 Excel 文件路径")
    parser.add_argument("--output", required=True, help="输出文件路径")
    parser.add_argument("--calculations", help="计算任务 JSON 字符串")
    parser.add_argument("--formatting", help="格式化配置 JSON 字符串")
    parser.add_argument("--charts", help="图表配置 JSON 字符串")
    parser.add_argument("--export-pdf", action="store_true", help="导出为 PDF")
    
    args = parser.parse_args()
    
    # 解析 JSON 参数
    calculations = json.loads(args.calculations) if args.calculations else None
    formatting = json.loads(args.formatting) if args.formatting else None
    charts = json.loads(args.charts) if args.charts else None
    
    # 执行处理
    result = process_excel_report(
        input_file=args.input,
        output_file=args.output,
        calculations=calculations,
        formatting=formatting,
        charts=charts,
        export_pdf=args.export_pdf
    )
    
    print(f"处理完成: {result}")
    return result


if __name__ == "__main__":
    main()
